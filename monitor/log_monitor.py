import re
import os
import openpyxl
from datetime import datetime
from collections import defaultdict
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines

def extract_metric(log_file_path, metric):
    data = []

    with open(log_file_path, 'r') as log_file:
        for line in log_file:
            # Ищем строки, содержащие заданную метрику
            if metric in line:
                # Извлекаем временную метку (с учетом квадратных скобок)
                timestamp_match = re.match(r'\[(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d+Z)', line)
                # Извлекаем значение заданной метрики
                metric_match = re.search(rf'{metric}=(\d+(\.\d+)?)', line)  # Поддержка дробных чисел

                if timestamp_match and metric_match:
                    # Преобразуем временную метку в формат, понятный Excel
                    time_str = timestamp_match.group(1)
                    try:
                        # Разбираем ISO формат и преобразуем в понятный Excel формат
                        parts = time_str.split('T')
                        date_part = parts[0]
                        time_part = parts[1].split('.')[0]  # Убираем микросекунды и Z
                        excel_time = f"{date_part} {time_part}"
                    except:
                        # Если что-то пошло не так, оставляем исходный формат
                        excel_time = time_str
                        
                    data.append([
                        excel_time,  # Время в более читаемом формате
                        float(metric_match.group(1))  # Значение заданной метрики как число
                    ])
    
    return data

def extract_slot_events(log_file_path):
    """
    Извлекает информацию о событиях для каждого слота:
    - new fork
    - replay-slot-stats
    - tower-vote latest
    """
    new_fork_events = {}
    replay_stats_events = {}
    tower_vote_events = {}
    tower_observed_root_diff = []  # Структура для хранения разницы tower-observed - root
    tower_vote_root_diff = []  # Новая структура для хранения разницы tower-vote - root
    slot_is_dead_data = []  # Новая структура для хранения is_dead статуса
    
    with open(log_file_path, 'r') as log_file:
        for line in log_file:
            # Извлекаем временную метку
            timestamp_match = re.match(r'\[(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d+Z)', line)
            
            if not timestamp_match:
                continue
                
            timestamp_str = timestamp_match.group(1)
            # Обрабатываем нестандартный формат микросекунд
            parts = timestamp_str.split('.')
            if len(parts) == 2:
                # Формат ISO с Z в конце
                date_part = parts[0]
                # Берем только первые 6 знаков микросекунд и удаляем Z в конце
                micro_part = parts[1][:-1]
                if len(micro_part) > 6:
                    micro_part = micro_part[:6]
                # Пересобираем строку
                timestamp_str = f"{date_part}.{micro_part}"
                timestamp = datetime.strptime(timestamp_str, '%Y-%m-%dT%H:%M:%S.%f')
            else:
                # Формат без микросекунд
                timestamp = datetime.strptime(timestamp_str.rstrip('Z'), '%Y-%m-%dT%H:%M:%S')
            
            # Форматируем время для Excel
            excel_time = timestamp.strftime('%Y-%m-%d %H:%M:%S')
            
            # Проверяем события для каждого слота
            # 1. new fork событие
            new_fork_match = re.search(r'new fork:(\d+)', line)
            if new_fork_match:
                slot = int(new_fork_match.group(1))
                new_fork_events[slot] = timestamp
            
            # 2. replay-slot-stats событие
            replay_stats_match = re.search(r'replay-slot-stats slot=(\d+)i', line)
            if replay_stats_match:
                slot = int(replay_stats_match.group(1))
                replay_stats_events[slot] = timestamp
            
            # 3. tower-vote latest событие
            tower_vote_match = re.search(r'tower-vote latest=(\d+)i', line)
            if tower_vote_match:
                slot = int(tower_vote_match.group(1))
                tower_vote_events[slot] = timestamp

            # 4. tower-observed и root
            if 'datapoint: tower-observed' in line:
                tower_observed_match = re.search(r'tower-observed slot=(\d+)i', line)
                root_match = re.search(r'root=(\d+)i', line)
                
                if tower_observed_match and root_match:
                    tower_observed = int(tower_observed_match.group(1))
                    root = int(root_match.group(1))
                    diff = tower_observed - root
                    
                    tower_observed_root_diff.append([
                        excel_time,         # Время
                        diff,               # Разница
                        tower_observed,     # tower-observed
                        root                # root
                    ])
                    
            # 5. tower-vote latest и root (НОВОЕ)
            if 'datapoint: tower-vote' in line and 'latest=' in line and 'root=' in line:
                tower_vote_match = re.search(r'tower-vote latest=(\d+)i', line)
                root_match = re.search(r'root=(\d+)i', line)
                
                if tower_vote_match and root_match:
                    tower_vote = int(tower_vote_match.group(1))
                    root = int(root_match.group(1))
                    diff = tower_vote - root
                    
                    tower_vote_root_diff.append([
                        excel_time,      # Время
                        diff,            # Разница
                        tower_vote,      # tower-vote latest
                        root             # root
                    ])
                    
            # 6. is_dead статус (НОВОЕ)
            if 'datapoint: slot_stats_tracking_complete' in line:
                slot_match = re.search(r'slot=(\d+)i', line)
                is_dead_match = re.search(r'is_dead=(true|false)', line)
                
                if slot_match and is_dead_match:
                    slot = int(slot_match.group(1))
                    is_dead = 1 if is_dead_match.group(1) == 'true' else 0  # Преобразуем в 0/1
                    
                    slot_is_dead_data.append([
                        excel_time,      # Время
                        is_dead,         # is_dead статус (0=false, 1=true)
                        slot             # Номер слота
                    ])
    
    return new_fork_events, replay_stats_events, tower_vote_events, tower_observed_root_diff, tower_vote_root_diff, slot_is_dead_data

def calculate_processing_times(new_fork_events, replay_stats_events, tower_vote_events):
    """
    Рассчитывает временные разницы между событиями для каждого слота
    """
    fork_to_replay_times = []  # Время от new fork до replay-slot-stats
    replay_to_vote_times = []  # Время от replay-slot-stats до tower-vote latest
    detailed_replay_vote_info = []  # Детальная информация для анализа выбросов
    
    # Для каждого слота с replay-slot-stats находим соответствующие события
    for slot, replay_time in replay_stats_events.items():
        # Время от new fork до replay-slot-stats
        if slot in new_fork_events:
            fork_time = new_fork_events[slot]
            time_diff_ms = (replay_time - fork_time).total_seconds() * 1000  # в миллисекундах
            # Добавляем форматированное время
            formatted_time = replay_time.strftime('%Y-%m-%d %H:%M:%S')
            fork_to_replay_times.append([slot, time_diff_ms, formatted_time])
        
        # Время от replay-slot-stats до tower-vote latest
        # Ищем ближайшее по времени событие tower-vote для этого слота
        found = False
        for vote_slot, vote_time in tower_vote_events.items():
            if vote_slot == slot and vote_time >= replay_time:
                time_diff_ms = (vote_time - replay_time).total_seconds() * 1000  # в миллисекундах
                # Добавляем форматированное время
                formatted_time = vote_time.strftime('%Y-%m-%d %H:%M:%S')
                replay_to_vote_times.append([slot, time_diff_ms, formatted_time])
                
                # Добавляем детальную информацию для анализа
                detailed_replay_vote_info.append({
                    'slot': slot,
                    'replay_time': replay_time.strftime('%Y-%m-%d %H:%M:%S.%f'),
                    'vote_time': vote_time.strftime('%Y-%m-%d %H:%M:%S.%f'),
                    'time_diff_ms': time_diff_ms
                })
                
                found = True
                break
    
    return fork_to_replay_times, replay_to_vote_times, detailed_replay_vote_info

def add_chart(worksheet, title=None):
    """
    Добавляет график на лист Excel с одной синей линией
    """
    chart = LineChart()
    if title:
        chart.title = title
    else:
        chart.title = worksheet.title
        
    chart.style = 2
    chart.y_axis.title = "Values"
    
    # Настройка оси Y с четкими метками шкалы
    from openpyxl.chart.axis import ChartLines
    
    # Правильно настраиваем линии сетки
    chart.y_axis.majorGridlines = ChartLines()  # Показываем горизонтальные линии сетки
    chart.y_axis.numFmt = "0.00"  # Формат чисел с двумя знаками после запятой
    chart.y_axis.scaling.min = 0  # Начинаем с нуля для лучшего понимания масштаба
    
    # Обеспечиваем отображение подписей вдоль оси
    chart.y_axis.delete = False
    chart.y_axis.tickLblPos = "nextTo"  # Располагаем метки рядом с осью
    
    # Если это график с задержками, настраиваем специальные интервалы
    if "delay" in str(title).lower() or "time" in str(title).lower():
        max_value = 0
        for row in range(2, worksheet.max_row + 1):
            val = worksheet.cell(row=row, column=2).value
            if val and val > max_value:
                max_value = val
        
        # Определяем подходящий интервал для делений оси
        if max_value > 1000:
            # Для больших значений (выбросы) используем больший интервал
            chart.y_axis.majorUnit = max_value / 10
        elif max_value > 100:
            chart.y_axis.majorUnit = 20
        elif max_value > 50:
            chart.y_axis.majorUnit = 10
        elif max_value > 10:
            chart.y_axis.majorUnit = 5
        else:
            chart.y_axis.majorUnit = 1
    
    # Отключаем легенду
    chart.legend = None
    
    # Проверяем, есть ли колонка временных меток
    has_timestamp = False
    if worksheet.max_column >= 3:
        header = worksheet.cell(row=1, column=3).value
        has_timestamp = header == 'timestamp'
    
    # Добавляем данные без заголовков
    data = Reference(worksheet, min_col=2, min_row=2, max_row=worksheet.max_row, max_col=2)
    chart.add_data(data, titles_from_data=False)
    
    # Используем временные метки для оси X, если они есть
    # Убираем заголовки осей
    chart.x_axis.title = None
    chart.y_axis.title = None
    
    # Используем временные метки или слоты как категории
    if has_timestamp:
        # Используем метки времени как категории
        cats = Reference(worksheet, min_col=3, min_row=2, max_row=worksheet.max_row)
        chart.set_categories(cats)
    else:
        # Используем слоты/индексы как категории
        cats = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
        chart.set_categories(cats)
    
    # Настраиваем тонкие линии синего цвета
    for series in chart.series:
        series.graphicalProperties.line.solidFill = "0000FF"  # Синий цвет
        series.graphicalProperties.line.width = 7200  # Минимальная толщина (0.75 pt)
        # Убираем все подписи данных на графике
        series.dLbls = None
    
    # Устанавливаем размер графика
    chart.height = 15  # высота в сантиметрах
    chart.width = 20   # ширина в сантиметрах
    
    # Добавляем график на лист
    worksheet.add_chart(chart, "E2")

def main():
    base_path = os.path.dirname(os.path.abspath(__file__))  # Относительный путь

    log_file_path = os.path.join(base_path, 'solana.log')  # Путь к лог-файлу
    metrics_file_path = os.path.join(base_path, 'metrics.txt')  # Путь к файлу с метриками
    output_excel_path = os.path.join(base_path, 'metrics.xlsx')  # Путь к выходному Excel файлу

    # Чтение списка метрик из файла
    with open(metrics_file_path, 'r') as metrics_file:
        metrics = [line.strip() for line in metrics_file if line.strip()]

    # Создаем новый Excel файл
    workbook = openpyxl.Workbook()

    # Извлекаем данные метрик
    for metric in metrics:
        data = extract_metric(log_file_path, metric)

        # Создаем новую вкладку для каждой метрики
        # Ограничиваем имя листа до 31 символа максимум, как требует Excel
        sheet_name = metric[:31]
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(['time', metric])  # Заголовки столбцов
        
        for entry in data:
            sheet.append(entry)  # Запись данных
            
        # Добавляем график, если есть данные
        if len(data) > 0:
            add_chart(sheet)

    # Установка формата для второго столбца (числа) начиная со второй строки
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in range(2, ws.max_row + 1):  # Начинаем со второй строки
            if ws.cell(row=row, column=2).value is not None:
                ws.cell(row=row, column=2).number_format = '0.00'  # Форматируем как число с двумя знаками после запятой

    # Анализ временных разниц между событиями и разниц tower-observed - root, tower-vote - root, is_dead
    new_fork_events, replay_stats_events, tower_vote_events, tower_observed_root_diff, tower_vote_root_diff, slot_is_dead_data = extract_slot_events(log_file_path)
    fork_to_replay_times, replay_to_vote_times, detailed_replay_vote_info = calculate_processing_times(
        new_fork_events, replay_stats_events, tower_vote_events
    )
    
    # Создаем новый лист для разницы tower-observed - root
    sheet_tower_observed_root = workbook.create_sheet(title="tower_observed_root_diff")
    sheet_tower_observed_root.append(['timestamp', 'diff', 'tower_observed', 'root'])  # Заголовки столбцов
    
    for entry in tower_observed_root_diff:
        sheet_tower_observed_root.append(entry)  # Запись данных
    
    # Добавляем график для разницы tower-observed - root
    if len(tower_observed_root_diff) > 0:
        add_chart(sheet_tower_observed_root, "Tower-Observed - Root (slots)")
        
    # Создаем новый лист для разницы tower-vote - root (НОВОЕ)
    sheet_tower_vote_root = workbook.create_sheet(title="tower_vote_root_diff")
    sheet_tower_vote_root.append(['timestamp', 'diff', 'tower_vote', 'root'])  # Заголовки столбцов
    
    for entry in tower_vote_root_diff:
        sheet_tower_vote_root.append(entry)  # Запись данных
    
    # Добавляем график для разницы tower-vote - root
    if len(tower_vote_root_diff) > 0:
        add_chart(sheet_tower_vote_root, "Tower-Vote - Root (slots)")
        
    # Создаем новый лист для is_dead статуса (НОВОЕ)
    sheet_is_dead = workbook.create_sheet(title="slot_is_dead")
    sheet_is_dead.append(['timestamp', 'is_dead', 'slot'])  # Заголовки столбцов
    
    for entry in slot_is_dead_data:
        sheet_is_dead.append(entry)  # Запись данных
    
    # Добавляем график для is_dead статуса
    if len(slot_is_dead_data) > 0:
        add_chart(sheet_is_dead, "Slot Is Dead Status (0=false, 1=true)")
        
    # Создаем лист для детального анализа задержек, если есть выбросы
    if detailed_replay_vote_info:
        times_ms = [info['time_diff_ms'] for info in detailed_replay_vote_info]
        avg_time = sum(times_ms) / len(times_ms)
        std_dev = (sum([(x - avg_time) ** 2 for x in times_ms]) / len(times_ms)) ** 0.5
        threshold = avg_time + 3 * std_dev  # 3 сигмы для выбросов
        
        # Создаем лист с детальной информацией о выбросах
        has_outliers = any(info['time_diff_ms'] > threshold for info in detailed_replay_vote_info)
        if has_outliers:
            sheet_outliers = workbook.create_sheet(title="vote_delay_outliers")
            sheet_outliers.append(['slot', 'replay_time', 'vote_time', 'delay_ms', 'threshold'])
            
            for info in detailed_replay_vote_info:
                if info['time_diff_ms'] > threshold:
                    sheet_outliers.append([
                        info['slot'],
                        info['replay_time'],
                        info['vote_time'],
                        info['time_diff_ms'],
                        threshold
                    ])
                    
                    print(f"Детальный анализ выброса для слота {info['slot']}:")
                    print(f"  Время replay: {info['replay_time']}")
                    print(f"  Время vote:   {info['vote_time']}")
                    print(f"  Задержка:     {info['time_diff_ms']:.2f} мс (порог: {threshold:.2f} мс)")
                    print("")
    
    # Создаем лист для времени между new fork и replay-slot-stats
    sheet_fork_to_replay = workbook.create_sheet(title="fork_to_replay_time")
    sheet_fork_to_replay.append(['slot', 'time_ms', 'timestamp'])  # Заголовки столбцов
    for entry in fork_to_replay_times:
        sheet_fork_to_replay.append(entry)
    
    # Создаем лист для времени между replay-slot-stats и tower-vote latest
    sheet_replay_to_vote = workbook.create_sheet(title="replay_to_vote_time")
    sheet_replay_to_vote.append(['slot', 'time_ms', 'timestamp', 'comment'])  # Заголовки столбцов
    
    # Находим среднее и стандартное отклонение для определения выбросов
    if replay_to_vote_times:
        times_ms = [entry[1] for entry in replay_to_vote_times]
        avg_time = sum(times_ms) / len(times_ms)
        std_dev = (sum([(x - avg_time) ** 2 for x in times_ms]) / len(times_ms)) ** 0.5
        threshold = avg_time + 3 * std_dev  # 3 сигмы для выбросов
        
        # Добавляем данные с комментариями о выбросах
        for entry in replay_to_vote_times:
            slot = entry[0]
            time_ms = entry[1]
            
            comment = ""
            if time_ms > threshold:
                comment = "Выброс! Возможные причины: высокая загрузка CPU, сетевые задержки, форк-выбор"
                print(f"Обнаружен выброс для слота {slot}: {time_ms} мс (порог: {threshold:.2f} мс)")
                
            sheet_replay_to_vote.append([slot, time_ms, entry[2], comment])
    else:
        # Если данных нет
        sheet_replay_to_vote.append([0, 0, "Нет данных"])
    
    # Добавляем графики для новых листов
    if len(fork_to_replay_times) > 0:
        add_chart(sheet_fork_to_replay, "Time from New Fork to Replay")
    
    if len(replay_to_vote_times) > 0:
        add_chart(sheet_replay_to_vote, "Time from Replay to Vote")

    # Удаление стандартного листа, если он пустой
    if 'Sheet' in workbook.sheetnames:
        std_sheet = workbook['Sheet']
        if std_sheet.max_row == 1:  # Если только заголовок
            workbook.remove(std_sheet)

    # Сохраняем Excel файл
    workbook.save(output_excel_path)
    print(f"Данные сохранены в {output_excel_path}")

# Запуск основной функции
if __name__ == "__main__":
    main()
